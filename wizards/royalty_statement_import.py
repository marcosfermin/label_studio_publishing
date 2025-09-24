# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
import csv
import io
import json
from datetime import datetime


class RoyaltyStatementImport(models.TransientModel):
    _name = 'royalty.statement.import'
    _description = 'Royalty Statement Import Wizard'

    # File Upload
    file_data = fields.Binary(string='Statement File', required=True,
                             help='CSV, TSV, or Excel file containing usage data')
    filename = fields.Char(string='File Name')
    
    # Source Configuration
    source_type = fields.Selection([
        ('distributor', 'Distributor'),
        ('pro', 'Performing Rights Organization'),
        ('publisher', 'Publisher'),
        ('youtube', 'YouTube Content ID'),
        ('spotify', 'Spotify for Artists'),
        ('apple', 'Apple Music for Artists'),
        ('custom', 'Custom Import')
    ], string='Source Type', required=True)
    
    source_id = fields.Many2one('res.partner', string='Source Partner')
    template_id = fields.Many2one('royalty.import.template', string='Import Template')
    
    # Period Configuration
    period_start = fields.Date(string='Period Start', required=True)
    period_end = fields.Date(string='Period End', required=True)
    reporting_date = fields.Date(string='Reporting Date', default=fields.Date.today)
    
    # Currency & Exchange
    currency_id = fields.Many2one('res.currency', string='Currency', required=True,
                                 default=lambda self: self.env.company.currency_id)
    exchange_rate = fields.Float(string='Exchange Rate', default=1.0, digits=(12, 6))
    
    # Processing Options
    dry_run = fields.Boolean(string='Dry Run (Preview Only)', default=True,
                            help='Preview import without creating records')
    auto_match = fields.Boolean(string='Auto-match Usage Lines', default=True,
                               help='Automatically attempt to match imported lines to catalog')
    skip_duplicates = fields.Boolean(string='Skip Duplicate Lines', default=True,
                                    help='Skip lines that appear to be duplicates based on key fields')
    
    # Batch Processing
    batch_size = fields.Integer(string='Batch Size', default=1000,
                               help='Number of lines to process at once')
    
    # File Analysis
    file_delimiter = fields.Selection([
        (',', 'Comma (,)'),
        (';', 'Semicolon (;)'),
        ('\\t', 'Tab'),
        ('|', 'Pipe (|)')
    ], string='Delimiter', default=',')
    
    has_header = fields.Boolean(string='File has Header Row', default=True)
    encoding = fields.Selection([
        ('utf-8', 'UTF-8'),
        ('latin-1', 'Latin-1'),
        ('cp1252', 'Windows-1252')
    ], string='File Encoding', default='utf-8')
    
    # Preview & Results
    preview_data = fields.Text(string='Preview Data', readonly=True)
    column_mapping = fields.Text(string='Column Mapping (JSON)', 
                                help='JSON mapping of file columns to usage line fields')
    
    # Progress Tracking
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('preview', 'Preview'),
        ('importing', 'Importing'),
        ('completed', 'Completed'),
        ('error', 'Error')
    ], string='State', default='draft')
    
    progress = fields.Float(string='Progress (%)', default=0.0)
    import_log = fields.Text(string='Import Log')
    
    # Results
    total_lines = fields.Integer(string='Total Lines', readonly=True)
    imported_lines = fields.Integer(string='Imported Lines', readonly=True)
    matched_lines = fields.Integer(string='Auto-matched Lines', readonly=True)
    error_lines = fields.Integer(string='Error Lines', readonly=True)
    import_batch_id = fields.Char(string='Import Batch ID', readonly=True)

    @api.onchange('source_type')
    def _onchange_source_type(self):
        """Load default template and mapping for source type"""
        if self.source_type:
            template = self.env['royalty.import.template'].search([
                ('source_type', '=', self.source_type),
                ('is_default', '=', True)
            ], limit=1)
            if template:
                self.template_id = template
                self.column_mapping = template.column_mapping
                self.file_delimiter = template.delimiter
                self.has_header = template.has_header

    @api.onchange('template_id')
    def _onchange_template(self):
        """Load template configuration"""
        if self.template_id:
            self.column_mapping = self.template_id.column_mapping
            self.file_delimiter = self.template_id.delimiter
            self.has_header = self.template_id.has_header

    def action_preview_file(self):
        """Preview the uploaded file and analyze its structure"""
        if not self.file_data:
            raise UserError(_('Please upload a file first'))
        
        try:
            # Decode file
            file_content = base64.b64decode(self.file_data)
            
            # Determine file type and parse
            if self.filename.endswith('.csv') or self.filename.endswith('.tsv'):
                preview_data = self._parse_csv_preview(file_content)
            elif self.filename.endswith(('.xls', '.xlsx')):
                preview_data = self._parse_excel_preview(file_content)
            else:
                raise UserError(_('Unsupported file format. Please upload CSV, TSV, or Excel files.'))
            
            self.preview_data = preview_data
            self.state = 'preview'
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'royalty.statement.import',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
            
        except Exception as e:
            raise UserError(_('Error parsing file: %s') % str(e))

    def action_configure_mapping(self):
        """Open column mapping wizard"""
        return {
            'name': _('Configure Column Mapping'),
            'type': 'ir.actions.act_window',
            'res_model': 'royalty.import.mapping.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_import_wizard_id': self.id,
                'default_preview_data': self.preview_data,
                'default_column_mapping': self.column_mapping,
            }
        }

    def action_import_statements(self):
        """Execute the import process"""
        if not self.file_data:
            raise UserError(_('Please upload a file first'))
        
        if not self.column_mapping:
            raise UserError(_('Please configure column mapping first'))
        
        try:
            self.state = 'importing'
            self.import_batch_id = self._generate_batch_id()
            
            # Parse column mapping
            mapping = json.loads(self.column_mapping)
            
            # Process file
            if self.dry_run:
                result = self._process_dry_run(mapping)
            else:
                result = self._process_import(mapping)
            
            # Update results
            self.total_lines = result.get('total_lines', 0)
            self.imported_lines = result.get('imported_lines', 0)
            self.matched_lines = result.get('matched_lines', 0)
            self.error_lines = result.get('error_lines', 0)
            self.import_log = result.get('import_log', '')
            self.progress = 100.0
            self.state = 'completed'
            
            return self._show_results()
            
        except Exception as e:
            self.state = 'error'
            self.import_log = f"Import failed: {str(e)}"
            raise UserError(_('Import failed: %s') % str(e))

    def action_view_imported_lines(self):
        """View imported usage lines"""
        if not self.import_batch_id:
            raise UserError(_('No import batch found'))
        
        return {
            'name': _('Imported Usage Lines'),
            'type': 'ir.actions.act_window',
            'res_model': 'royalty.usage.line',
            'view_mode': 'tree,form',
            'domain': [('import_batch_id', '=', self.import_batch_id)],
            'context': {'search_default_group_matched_state': 1}
        }

    def _parse_csv_preview(self, file_content):
        """Parse CSV content and return preview"""
        try:
            content_str = file_content.decode(self.encoding)
            delimiter = '\\t' if self.file_delimiter == '\\t' else self.file_delimiter
            
            csv_reader = csv.reader(io.StringIO(content_str), delimiter=delimiter)
            lines = list(csv_reader)
            
            if not lines:
                raise UserError(_('File appears to be empty'))
            
            # Get header and sample rows
            header = lines[0] if self.has_header else None
            sample_rows = lines[1:6] if self.has_header else lines[:5]
            
            preview = {
                'header': header,
                'sample_rows': sample_rows,
                'total_rows': len(lines) - (1 if self.has_header else 0)
            }
            
            return json.dumps(preview, indent=2)
            
        except Exception as e:
            raise UserError(_('Error parsing CSV: %s') % str(e))

    def _parse_excel_preview(self, file_content):
        """Parse Excel content and return preview"""
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            workbook = load_workbook(io.BytesIO(file_content))
            sheet = workbook.active
            
            # Get header and sample rows
            rows = list(sheet.iter_rows(values_only=True))
            header = list(rows[0]) if self.has_header and rows else None
            sample_rows = [list(row) for row in rows[1:6]] if self.has_header else [list(row) for row in rows[:5]]
            
            preview = {
                'header': header,
                'sample_rows': sample_rows,
                'total_rows': len(rows) - (1 if self.has_header else 0)
            }
            
            return json.dumps(preview, indent=2)
            
        except ImportError:
            raise UserError(_('openpyxl library not installed. Cannot process Excel files.'))
        except Exception as e:
            raise UserError(_('Error parsing Excel: %s') % str(e))

    def _process_dry_run(self, mapping):
        """Process file in dry-run mode for validation"""
        errors = []
        matched_count = 0
        total_count = 0
        
        # Parse file content
        usage_data = self._parse_file_content(mapping)
        
        for line_data in usage_data:
            total_count += 1
            
            # Validate required fields
            validation_errors = self._validate_usage_line(line_data)
            if validation_errors:
                errors.extend(validation_errors)
                continue
            
            # Test auto-matching if enabled
            if self.auto_match:
                confidence = self._test_matching(line_data)
                if confidence > 0.7:  # Configurable threshold
                    matched_count += 1
        
        return {
            'total_lines': total_count,
            'imported_lines': 0,  # Dry run doesn't import
            'matched_lines': matched_count,
            'error_lines': len(errors),
            'import_log': f"Dry run completed.\
Potential matches: {matched_count}\
Errors: {len(errors)}\
" + "\
".join(errors[:10])
        }

    def _process_import(self, mapping):
        """Process actual import"""
        imported_count = 0
        matched_count = 0
        error_count = 0
        errors = []
        
        # Parse file content
        usage_data = self._parse_file_content(mapping)
        total_count = len(usage_data)
        
        # Process in batches
        batch_count = 0
        batch_data = []
        
        for i, line_data in enumerate(usage_data):
            batch_data.append(line_data)
            
            if len(batch_data) >= self.batch_size or i == total_count - 1:
                # Process batch
                batch_result = self._process_batch(batch_data)
                imported_count += batch_result['imported']
                matched_count += batch_result['matched']
                error_count += batch_result['errors']
                errors.extend(batch_result['error_messages'])
                
                # Update progress
                self.progress = (i + 1) / total_count * 100
                self.env.cr.commit()  # Commit progress
                
                batch_data = []
                batch_count += 1
        
        return {
            'total_lines': total_count,
            'imported_lines': imported_count,
            'matched_lines': matched_count,
            'error_lines': error_count,
            'import_log': f"Import completed.\
Processed {batch_count} batches\
Errors: {error_count}\
" + "\
".join(errors[:10])
        }

    def _process_batch(self, batch_data):
        """Process a batch of usage lines"""
        imported = 0
        matched = 0
        errors = 0
        error_messages = []
        
        usage_lines = []
        
        for line_data in batch_data:
            try:
                # Validate
                validation_errors = self._validate_usage_line(line_data)
                if validation_errors:
                    errors += 1
                    error_messages.extend(validation_errors)
                    continue
                
                # Check for duplicates if enabled
                if self.skip_duplicates and self._is_duplicate(line_data):
                    continue
                
                # Create usage line
                usage_line_vals = self._prepare_usage_line_vals(line_data)
                usage_lines.append(usage_line_vals)
                imported += 1
                
            except Exception as e:
                errors += 1
                error_messages.append(f"Line error: {str(e)}")
        
        # Bulk create
        if usage_lines:
            created_lines = self.env['royalty.usage.line'].create(usage_lines)
            
            # Auto-match if enabled
            if self.auto_match:
                for line in created_lines:
                    line.action_auto_match()
                    if line.matched_state in ['auto_matched', 'manually_matched']:
                        matched += 1
        
        return {
            'imported': imported,
            'matched': matched,
            'errors': errors,
            'error_messages': error_messages
        }

    def _parse_file_content(self, mapping):
        """Parse file content based on mapping"""
        file_content = base64.b64decode(self.file_data)
        
        if self.filename.endswith('.csv') or self.filename.endswith('.tsv'):
            return self._parse_csv_data(file_content, mapping)
        elif self.filename.endswith(('.xls', '.xlsx')):
            return self._parse_excel_data(file_content, mapping)
        else:
            raise UserError(_('Unsupported file format'))

    def _parse_csv_data(self, file_content, mapping):
        """Parse CSV data with column mapping"""
        content_str = file_content.decode(self.encoding)
        delimiter = '\\t' if self.file_delimiter == '\\t' else self.file_delimiter
        
        csv_reader = csv.DictReader(io.StringIO(content_str), delimiter=delimiter) if self.has_header else csv.reader(io.StringIO(content_str), delimiter=delimiter)
        
        usage_data = []
        for row in csv_reader:
            if self.has_header:
                # Map columns by name
                line_data = {}
                for field, column in mapping.items():
                    if column in row:
                        line_data[field] = row[column]
            else:
                # Map columns by index
                line_data = {}
                for field, column_index in mapping.items():
                    if isinstance(column_index, int) and column_index < len(row):
                        line_data[field] = row[column_index]
            
            usage_data.append(line_data)
        
        return usage_data

    def _prepare_usage_line_vals(self, line_data):
        """Prepare values for creating usage line record"""
        vals = {
            'source_type': self.source_type,
            'source_id': self.source_id.id if self.source_id else False,
            'period_start': self.period_start,
            'period_end': self.period_end,
            'reporting_date': self.reporting_date,
            'currency_id': self.currency_id.id,
            'exchange_rate': self.exchange_rate,
            'import_batch_id': self.import_batch_id,
        }
        
        # Map line data to fields
        field_mapping = {
            'track_name': 'track_name',
            'artist_name': 'artist_name',
            'album_name': 'album_name',
            'isrc': 'isrc',
            'iswc': 'iswc',
            'upc': 'upc',
            'usage_type': 'usage_type',
            'service': 'service',
            'territory_code': 'territory_code',
            'units': 'units',
            'gross_amount': 'gross_amount',
            'fees': 'fees',
        }
        
        for field, line_key in field_mapping.items():
            if line_key in line_data and line_data[line_key]:
                value = line_data[line_key]
                
                # Type conversion
                if field in ['units']:
                    try:
                        vals[field] = int(value)
                    except (ValueError, TypeError):
                        vals[field] = 0
                elif field in ['gross_amount', 'fees']:
                    try:
                        vals[field] = float(value)
                    except (ValueError, TypeError):
                        vals[field] = 0.0
                else:
                    vals[field] = str(value).strip()
        
        return vals

    def _validate_usage_line(self, line_data):
        """Validate a single usage line data"""
        errors = []
        
        # Required fields validation
        required_fields = ['track_name', 'artist_name']
        for field in required_fields:
            if not line_data.get(field):
                errors.append(f"Missing required field: {field}")
        
        # Data type validation
        numeric_fields = ['units', 'gross_amount', 'fees']
        for field in numeric_fields:
            if field in line_data and line_data[field]:
                try:
                    float(line_data[field])
                except (ValueError, TypeError):
                    errors.append(f"Invalid numeric value for {field}: {line_data[field]}")
        
        return errors

    def _test_matching(self, line_data):
        """Test matching confidence for a line"""
        # Simplified matching test - in production, use the actual matching engine
        if line_data.get('isrc'):
            # ISRC match would have high confidence
            recording = self.env['music.recording'].search([('isrc', '=', line_data['isrc'])], limit=1)
            return 1.0 if recording else 0.0
        
        if line_data.get('track_name') and line_data.get('artist_name'):
            # Fuzzy title/artist match
            recordings = self.env['music.recording'].search([
                ('title', 'ilike', line_data['track_name']),
                ('main_artist_ids.name', 'ilike', line_data['artist_name'])
            ], limit=1)
            return 0.8 if recordings else 0.0
        
        return 0.0

    def _is_duplicate(self, line_data):
        """Check if this line is a duplicate"""
        # Simple duplicate check based on key fields
        domain = [
            ('source_type', '=', self.source_type),
            ('period_start', '=', self.period_start),
            ('period_end', '=', self.period_end),
        ]
        
        if line_data.get('track_name'):
            domain.append(('track_name', '=', line_data['track_name']))
        if line_data.get('artist_name'):
            domain.append(('artist_name', '=', line_data['artist_name']))
        if line_data.get('isrc'):
            domain.append(('isrc', '=', line_data['isrc']))
        
        return bool(self.env['royalty.usage.line'].search(domain, limit=1))

    def _generate_batch_id(self):
        """Generate unique batch ID"""
        return f"{self.source_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{self.id}"

    def _show_results(self):
        """Show import results"""
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'royalty.statement.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }


class RoyaltyImportTemplate(models.Model):
    _name = 'royalty.import.template'
    _description = 'Royalty Import Template'
    
    name = fields.Char(string='Template Name', required=True)
    source_type = fields.Selection([
        ('distributor', 'Distributor'),
        ('pro', 'Performing Rights Organization'),
        ('publisher', 'Publisher'),
        ('youtube', 'YouTube Content ID'),
        ('spotify', 'Spotify for Artists'),
        ('apple', 'Apple Music for Artists'),
        ('custom', 'Custom Import')
    ], string='Source Type', required=True)
    
    is_default = fields.Boolean(string='Default Template')
    column_mapping = fields.Text(string='Column Mapping (JSON)', required=True)
    delimiter = fields.Selection([
        (',', 'Comma (,)'),
        (';', 'Semicolon (;)'),
        ('\\t', 'Tab'),
        ('|', 'Pipe (|)')
    ], string='Delimiter', default=',')
    has_header = fields.Boolean(string='Has Header Row', default=True)
    notes = fields.Text(string='Notes')
    
    active = fields.Boolean(string='Active', default=True)